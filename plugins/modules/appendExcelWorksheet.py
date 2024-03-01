#s.fauquembergue(sii)
#v1.0
#initial release 30/10/2023

from ansible.module_utils.basic import AnsibleModule
import os, re
from openpyxl import Workbook, load_workbook

def main():
    #region params
    module_args = dict(
        hostname = dict(required=True, type='str'),
        workbook = dict(required=True, type='str'),
        worksheet = dict(required=True, type='str'),
        datas = dict(required=False, type='list'),
        )

    #endregion

    #region ansible
    module = AnsibleModule(
        argument_spec=module_args,
        supports_check_mode=True
    )

    hostname = module.params.get('hostname')
    workbook = module.params.get('workbook')
    worksheet = module.params.get('worksheet')
    datas = module.params.get('datas')

    result = dict(
        changed=False,
        skipped=True,
        message=None
    )
    #endregion

    #region get
    output=None
    wb=load_workbook(workbook)
    ws=wb[worksheet]

    output=worksheet
    result["message"]=output
    result["skipped"]=False
    #endregion

    #region test
    compliant=True

    line=2
    matched=False

    while ws['A'+str(line)].value and matched==False:
        if ws['A'+str(line)].value == hostname:
            matched=True

        line=line+1

    if matched:
        result["message"]='Values from "'+hostname+'" are in file.'

    else:
       compliant=False

    if module.check_mode:
        wb.close()
        module.exit_json(**result)
    #endregion

    #region set
    if not compliant:
        ##write datas to worksheet
        line = 2
        countAddedLine=0

        while ws['A'+str(line)].value:
           line=line+1

        for d in datas:
            ws['A'+str(line)]=hostname

            if len(hostname)>ws.column_dimensions['A'].width:
                ws.column_dimensions['A'].width=len(hostname)+1

            asciiPosition=ord('B')

            for d1 in d:
                cell=chr(asciiPosition)+str(line)
                cellValue=d[d1]

                ws[cell]=cellValue

                if len(str(cellValue))>ws.column_dimensions[chr(asciiPosition)].width:
                    ws.column_dimensions[chr(asciiPosition)].width=len(str(cellValue))+1

                asciiPosition=asciiPosition+1

            line=line+1
            countAddedLine=countAddedLine+1

        wb.save(workbook)
        result["message"]=str(countAddedLine)+" line(s) added"
        result["changed"]=True
    #endregion

    wb.close()
    module.exit_json(**result)

if __name__ == '__main__':
    main()

