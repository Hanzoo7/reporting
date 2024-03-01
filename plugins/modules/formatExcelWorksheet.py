#s.fauquembergue(sii)
#v1.0
#initial release 31/10/2023

from ansible.module_utils.basic import AnsibleModule
import os, re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import colors, Color, PatternFill, Font, Border

def main():
    #region params
    module_args = dict(
        workbook = dict(required=True, type='str'),
        worksheet = dict(required=True, type='str'),
        formatHeaders = dict(required=False, type='bool'),
        freezeRows = dict(required=False, type='int'),
        headColor = dict(required=False, type='str'),
        )
    #endregion

    #region ansible
    module = AnsibleModule(
        argument_spec=module_args,
        supports_check_mode=True
    )

    workbook = module.params.get('workbook')
    worksheet = module.params.get('worksheet')
    formatHeaders = module.params.get('formatHeaders')
    freezeRows = module.params.get('freezeRows')
    headColor = module.params.get('headColor')

    result = dict(
        changed=False,
        skipped=True,
        message=None
    )
    #endregion

    #region get
    output=None
    wb=load_workbook(workbook)
    ws=wb[worksheet]

    output=worksheet
    result["message"]=[formatHeaders,freezeRows,headColor]
    result["skipped"]=False
    #endregion

    #region test
    compliant=True

    if formatHeaders and not freezeRows and ws.freeze_panes!="A2":
        compliant=False

    if formatHeaders and  freezeRows and ws.freeze_panes!=chr(65+freezeRows)+"2":
        compliant=False

    if module.check_mode:
        wb.close()
        module.exit_json(**result)
    #endregion

    #region set
    if not compliant:
        if formatHeaders and headColor:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref=ws.dimensions
            index=0

            headStyle=PatternFill(
                start_color=headColor,
                end_color=headColor,
                fill_type='solid')

            for c in ws.columns:
                ws[chr(65+index)+"1"].fill=headStyle
                index=index+1

        if freezeRows:
            ws.freeze_panes = chr(65+freezeRows)+"2"

        wb.save(workbook)
        result["message"]=output
        result["changed"]=True
    #endregion

    wb.close()
    module.exit_json(**result)

if __name__ == '__main__':
    main()


