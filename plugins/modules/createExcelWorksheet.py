#s.fauquembergue(sii)
#v1.0
#initial release 26/10/2023

from ansible.module_utils.basic import AnsibleModule
import re, ast
from openpyxl import Workbook, load_workbook

def main():
    #region params
    module_args = dict(
        workbook = dict(required=True, type='str'),
        worksheet = dict(required=False, type='str'),
        datas = dict(required=False, type='list'),
        )

    #endregion

    #region ansible
    module = AnsibleModule(
        argument_spec=module_args,
        supports_check_mode=True
    )

    workbook = module.params.get('workbook')
    worksheet = module.params.get('worksheet')
    datas = module.params.get('datas')

    result = dict(
        changed=False,
        skipped=True,
        message=None
    )
    #endregion

    #region get
    output={}
    wb=load_workbook(workbook)
    wss=wb.sheetnames

    output["worksheet"]=worksheet
    result["message"]=output
    result["skipped"]=False
    #endregion

    #region test
    compliant=True

    if worksheet not in wss:
        compliant=False

    if module.check_mode:
        wb.close()
        module.exit_json(**result)
    #endregion

    #region set
    if not compliant:
        ws=wb.create_sheet(worksheet)
        output["columns"]=[]
        fields={}

        if datas:
            fields=datas[0]

        ws['A1'] = 'hostname'
        output["columns"].append('hostname')
        asciiPosition=ord('B')

        for f in fields:
            ws[chr(asciiPosition)+str('1')]=f
            asciiPosition = asciiPosition+1
            output["columns"].append(f)

        wb.save(workbook)
        result["message"]=output
        result["changed"]=True
    #endregion

    wb.close()
    module.exit_json(**result)

if __name__ == '__main__':
    main()

